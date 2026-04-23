#!/usr/bin/env python3
"""Build vk-resources-inventory.xlsx — team audit / gap-analysis spreadsheet.

Now organized by Template family (T1-T5) per the build-playbook.

Columns:
  Resource | Template | Header | Status | Live link | Define | Comments |
  Additional Link 1 | Additional Link 2 | Vintageking.com equivalent

Tabs (template-first):
  T1 Editorial Light  · T2 Deep Article  · T3 Program / Service
  T4 Promo / Drop     · T5 Policy / Long-form
  PLP variants        · Core ecommerce (locked)  · Reference & tools
  Archive
"""
from __future__ import annotations
import re, html
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
LIVE = "https://barreletics.github.io/vk-secondary-pages"
CORE = "https://barreletics.github.io/vintage-king-redesign"

# ---------- helpers ----------
TITLE_RE = re.compile(r"<title>(.*?)</title>", re.S | re.I)
META_DESC_RE = re.compile(
    r'<meta[^>]+name=["\']description["\'][^>]+content=["\']([^"\']+)["\']', re.I
)
H1_RE = re.compile(r"<h1[^>]*>(.*?)</h1>", re.S | re.I)
TAG_RE = re.compile(r"<[^>]+>")

def clean(s: str) -> str:
    s = TAG_RE.sub(" ", s or "")
    return html.unescape(re.sub(r"\s+", " ", s)).strip()

def extract(path: Path) -> dict:
    if not path.exists():
        return {"title": "", "desc": "", "h1": ""}
    txt = path.read_text(encoding="utf-8", errors="ignore")
    title = clean((TITLE_RE.search(txt) or [None, ""])[1] if TITLE_RE.search(txt) else "")
    desc = (META_DESC_RE.search(txt) or [None, ""])[1] if META_DESC_RE.search(txt) else ""
    h1 = clean((H1_RE.search(txt) or [None, ""])[1] if H1_RE.search(txt) else "")
    return {"title": title, "desc": clean(desc), "h1": h1}

def url(rel: str) -> str:
    return f"{LIVE}/{rel}"

def define_for(meta: dict, fallback: str = "") -> str:
    return meta["desc"] or meta["h1"] or meta["title"] or fallback

# ---------- row builder (10 columns) ----------
def row(name, template, header, status, link, define,
        comments="", add1=("",""), add2=("",""), vk_equiv=""):
    return [name, template, header, status, link, define, comments,
            add1[1] if add1[0] else "",
            add2[1] if add2[0] else "",
            vk_equiv]

def labeled_link(label, url_):
    return (label, f"{label} — {url_}") if url_ else ("", "")

def status_for(rel: str) -> str:
    p = ROOT / rel
    if not p.exists():
        return "TBD"
    txt = p.read_text(encoding="utf-8", errors="ignore")
    if "PLACEHOLDER" in txt or len(txt) < 1500:
        return "Placeholder"
    return "Built"

# ---------- template-grouped page items ----------
# Each item: (display_name, base_filename, has_v2, vk_path, header_used)
T1_ITEMS = [  # Editorial Light — image-led marketing landings
    ("About VK",                      "about",                       True,  "/about",        "H2"),
    ("Make Your Mark",                "make-your-mark",              False, "",              "H2"),
    ("PLAYBACK Magazine",             "playback",                    False, "",              "H2"),
    ("Careers",                       "careers",                     False, "/careers",      "H2"),
    ("25 Years of Pro Audio",         "25-years-of-pro-audio",       True,  "",              "H2"),
    ("Live Sound How-To",             "live-sound-how",              False, "",              "H2"),
    ("Recording at Home",             "recording-at-home",           False, "",              "H2"),
    ("Classic Studios Gone Modern",   "classic-studios-gone-modern", False, "",              "H2"),
    ("Multi-Room Recording Studios",  "multi-room-recording-studios",True,  "",              "H2"),
]
T2_ITEMS = [  # Deep Article — long-form heritage / HOF
    ("Hall of Fame (index)",          "hall-of-fame",                False, "",              "H3"),
    ("HOF — Neve 1073 (overview)",    "neve-1073",                   False, "",              "H3"),
    ("HOF — Neve 1073 Buyer's Guide", "neve-1073-guide",             True,  "",              "H3"),
    ("HOF — Neve 1081",               "neve-1081",                   True,  "",              "H3"),
    ("HOF — Neumann U47",             "neumann-u47",                 False, "",              "H3"),
    ("HOF — Neumann U67",             "neumann-u67",                 False, "",              "H3"),
    ("HOF — Neumann U47 FET",         "neumann-u47-fet",             True,  "",              "H3"),
    ("HOF — Neumann Mic Comparison",  "neumann-mic-comparison",      False, "",              "H3"),
    ("HOF — Telefunken ELA-M 251",    "telefunken-ela-m-251",        True,  "",              "H3"),
    ("HOF — TAB Telefunken V72/V76",  "tab-telefunken-v72-v76",      True,  "",              "H3"),
    ("HOF — Coles 4038",              "coles-4038",                  True,  "",              "H3"),
    ("HOF — Fairchild 660/670",       "fairchild-660-670",           False, "",              "H3"),
    ("HOF — UREI 1176",               "urei-1176",                   False, "",              "H3"),
    ("HOF — LA-2A",                   "la-2a",                       False, "",              "H3"),
    ("HOF — Universal Audio 175B",    "universal-audio-175b",        True,  "",              "H3"),
    ("HOF — Universal Audio Apollo X","universal-audio-apollo-x",    True,  "",              "H3"),
]
T3_ITEMS = [  # Program / Service — process + pricing
    ("Studio Professionals",          "studio-professionals",        False, "",              "H2"),
    ("Trade Program",                 "trade-program",               False, "",              "H2"),
    ("Audio Consultants",             "audio-consultants",           False, "",              "H2"),
    ("Section 179 Tax Savings",       "section-179",                 False, "",              "H2"),
    ("VK Warranty",                   "warranty",                    False, "/warranty",     "H2"),
    ("VK Credit Card",                "vk-credit-card",              False, "",              "H2"),
]
T4_ITEMS = [  # Promo / Drop — time-bound campaign landings
    ("Back to School",                "back-to-school",              True,  "",              "H4"),
    ("Black Friday Microphone Deals", "black-friday-microphone-deals",True, "",              "H4"),
    ("New at NAMM — Microphones",     "new-at-namm-microphones",     True,  "",              "H4"),
    ("Universal Audio Promotions",    "universal-audio-promotions",  True,  "",              "H4"),
    ("Avid Pro Tools Trade-In",       "avid-pro-tools-trade-in",     True,  "",              "H4"),
    ("Avid S4 Control Surface Configs","avid-s4-control-surface-configurations", True, "",   "H4"),
    ("Avid S6 Control Surface Configs","avid-s6-control-surface-configurations", True, "",   "H4"),
]
T5_ITEMS = [  # Policy / Long-form — utility, no hero photo
    ("Privacy Policy",                "privacy-policy",              True,  "/policies/privacy-policy",  "H2"),
    ("Shipping Policy",               "shipping-policy",             True,  "/policies/shipping-policy", "H2"),
    ("Returns",                       "returns",                     True,  "/policies/refund-policy",   "H2"),
]

def render_items(items, template_label):
    rows = []
    for name, base, has_v2, vk_path, header in items:
        v2_file = f"{base}-v2.html"
        v1_file = f"{base}.html"
        primary = v2_file if has_v2 and (ROOT / v2_file).exists() else v1_file
        meta = extract(ROOT / primary)
        define = define_for(meta, name)
        comments = ""
        add1 = ("", "")
        if has_v2 and (ROOT / v1_file).exists() and primary == v2_file:
            add1 = labeled_link("v1 (older variant)", url(v1_file))
            comments = "v2 is current; v1 kept for reference and may move to /archive/."
        vk_equiv = f"https://www.vintageking.com{vk_path}" if vk_path else ""
        rows.append(row(name, template_label, header, status_for(primary),
                        url(primary), define, comments, add1, ("",""), vk_equiv))
    return rows

# ---------- non-template tabs ----------
def tab_plp_variants():
    rows = []
    items = [
        ("Microphones — PLP Default",            "microphones-plp-default.html",
         "C1", "H1",
         "Dark band hero, 3-up product grid, sticky filters. Default for all category landings.",
         "Recommended default.",
         "https://www.vintageking.com/collections/microphones"),
        ("Microphones — PLP Spotlight (large)",  "microphones-plp-spotlight.html",
         "C2", "H1",
         "Split hero (image right) + 3-up grid. Marquee categories with strong hero photo.",
         "Under team A/B vs Spotlight Compact.", ""),
        ("Microphones — PLP Spotlight Compact",  "microphones-plp-spotlight-compact.html",
         "C2", "H1",
         "Same as Spotlight — hero shrunk to ~440px so pill filters land above the fold.",
         "Under team A/B vs Spotlight large.", ""),
        ("Microphones — PLP Editorial",          "microphones-plp-editorial.html",
         "C3", "H1",
         "6-grid with feature tile + dark editorial interruption. Story-led collections.", "", ""),
        ("Microphones — PLP Spotlight Compact (font lab)", "microphones-plp-spotlight-compact-font-lab.html",
         "C2", "H1",
         "Font-comparison variant — used by the type-test workflow.",
         "Decision tool — not for production.", ""),
        ("Font Comparison PLP",                  "font-comparison-plp.html",
         "C2", "H1",
         "Side-by-side font comparison on a category PLP.",
         "Decision tool — feeds the type-test page in vk-redesign-final.", ""),
        ("Header Retail Three (lab)",            "header-retail-three.html",
         "—", "H1",
         "Header lab — three retail-leaning header variants on a PLP.", "Lab only.", ""),
        ("Hero Retail Three (lab)",              "hero-retail-three.html",
         "—", "H1",
         "Hero lab — three retail-leaning hero variants.", "Lab only.", ""),
    ]
    for name, rel, cat, header, define, comments, vk in items:
        rows.append(row(name, cat, header, status_for(rel),
                        url(rel), define, comments, ("",""), ("",""), vk))
    return rows

def tab_core_ecommerce():
    rows = []
    rows.append(row(
        "Core single-file prototype (latest)", "—", "—", "Locked",
        f"{CORE}/VintageKing-Redesign-v2_67-Mar2026.html",
        "Full single-file VK redesign prototype — all 10 core ecommerce screens.",
        "Source of truth for core ecommerce. DO NOT EDIT from this repo.",
        labeled_link("Mega-menu demo", f"{CORE}/mega-menu/VintageKing-MegaMenu-CursorLab-v2.html"),
        labeled_link("Filter / PLP demo (v4)", f"{CORE}/mega-menu/VK-Microphones-CategoryDemo-v4-refined.html"),
        "https://www.vintageking.com",
    ))
    pages = [
        ("Homepage",            "pages/home.html",     "https://www.vintageking.com"),
        ("Category page",       "pages/category.html", "https://www.vintageking.com/collections/microphones"),
        ("Product page",        "pages/product.html",  "https://www.vintageking.com/products/"),
        ("FAQ",                 "pages/faq.html",      "https://www.vintageking.com/pages/faq"),
        ("Demo Deals",          "pages/deals.html",    "https://www.vintageking.com/collections/demo-deals"),
        ("Sell or Trade",       "pages/sell.html",     "https://www.vintageking.com/sell-or-trade"),
        ("Studio Installations","pages/studio.html",   "https://www.vintageking.com/studio-design"),
        ("Tech Shop",           "pages/techshop.html", "https://www.vintageking.com/tech-shop"),
        ("Financing",           "pages/financing.html","https://www.vintageking.com/financing"),
        ("Open Box",            "pages/openbox.html",  "https://www.vintageking.com/collections/open-box"),
    ]
    for name, rel, vk in pages:
        meta = extract(ROOT / rel)
        rows.append(row(name, "—", "—", status_for(rel),
                        url(rel), define_for(meta, name),
                        "Standalone extract from core prototype.",
                        ("",""), ("",""), vk))
    return rows

def tab_reference_tools():
    rows = []
    items = [
        ("VK Build Playbook (READ FIRST)",
         "https://barreletics.github.io/vk-redesign-final/build-playbook.html",
         "Single source of truth: rules, rhythm, templates, headers, layouts, strips, type, color, SEO, handoff. Read this before touching any page.",
         "Lives in vk-redesign-final repo."),
        ("VK Redesign Final — Dashboard",
         "https://barreletics.github.io/vk-redesign-final/",
         "Roadmap dashboard. Read · See · Decide · Hand off.", ""),
        ("Type Test (decision tool)",
         "https://barreletics.github.io/vk-redesign-final/type-test.html",
         "A/B Editorial Playfair vs Retail DM Sans on product names across 3 surfaces.",
         "Vote here to lock global typography."),
        ("Value Strip Picker (decision tool)",
         "https://barreletics.github.io/vk-redesign-final/value-strips/value-strip-options.html",
         "11 value-strip variants side-by-side. Pick keepers; flag kills.", ""),
        ("Section Buffet (decision tool)",
         "https://barreletics.github.io/vk-redesign-final/section-buffet.html",
         "Every body section with Keep / Maybe / Kill markers.", ""),
        ("All Pages navigator (old repo inventory)",
         url("navigator.html"),
         "Master visual inventory of every page in the source repo.", ""),
        ("Style Guide (slim)",
         "https://barreletics.github.io/vk-redesign-final/style-guide.html",
         "Slim spec: colors, type scale, spacing, no-go list.",
         "Deeper detail still in old style-guide.html below."),
        ("Style Guide (deep, old repo)",
         url("style-guide.html"),
         "Full design system: tokens, badges, buttons, JSON-LD library, strip patterns.", ""),
        ("Page Layouts catalog",
         "https://barreletics.github.io/vk-redesign-final/page-layouts.html",
         "Visual catalog of every layout pattern.", ""),
        ("Source pages (combined)",
         url("VK-secondary-pages-source.html"),
         "Combined source containing copy, hero patterns, JSON-LD blocks for all secondary pages.", ""),
    ]
    for name, link, define, comments in items:
        rows.append(row(name, "—", "—", "Reference", link, define, comments))
    repo_base = "https://github.com/barreletics/vk-secondary-pages/blob/main"
    md_items = [
        ("HANDOFF.md — repo brief",      f"{repo_base}/HANDOFF.md",
         "Project guardrails, file inventory, page templates, build rules."),
        ("CLAUDE.md — design system rules", f"{repo_base}/CLAUDE.md",
         "Full design system briefing for AI agents."),
        ("README.md",                    f"{repo_base}/README.md",
         "How to open the navigator locally."),
    ]
    for name, link, define in md_items:
        rows.append(row(name, "—", "—", "Reference", link, define, ""))
    return rows

def tab_archive():
    rows = []
    rows.append(row(
        "(folder not yet created — pending buffet reorg Phase 4)",
        "—", "—", "Planned", "",
        "v1 duplicates and superseded files will move to /archive/ once the new dashboard ships.",
        "No action required from team — internal cleanup.",
    ))
    v1_candidates = sorted([p.name for p in ROOT.glob("*.html")
                            if not p.name.endswith("-v2.html")
                            and (ROOT / p.name.replace(".html", "-v2.html")).exists()])
    for v1 in v1_candidates:
        meta = extract(ROOT / v1)
        rows.append(row(
            f"[archive candidate] {v1}",
            "—", "—", "Archive candidate",
            url(v1), define_for(meta, v1),
            "Older v1 — superseded by v2 of same name.",
            labeled_link("Current v2", url(v1.replace(".html", "-v2.html"))),
        ))
    rows.append(row(
        "[archive candidate] templates-guide.html",
        "—", "—", "Archive candidate",
        url("templates-guide.html"),
        "Old agency template guide. Superseded by build-playbook.html in vk-redesign-final.",
        "Will move to /archive/ once team confirms build-playbook coverage.",
        labeled_link("Replacement", "https://barreletics.github.io/vk-redesign-final/build-playbook.html"),
    ))
    return rows

# ---------- workbook ----------
HEADERS = [
    "Resource",
    "Template",
    "Header",
    "Status",
    "Live link (our prototype)",
    "Define resource",
    "Comments",
    "Additional link 1",
    "Additional link 2",
    "Vintageking.com equivalent (team to confirm)",
]
COL_WIDTHS = [42, 14, 10, 16, 60, 60, 44, 50, 50, 44]

def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1A1A18")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    border = Border(*([Side(style="thin", color="DDDDDD")] * 4))
    align = Alignment(vertical="top", wrap_text=True)
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="left")
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS[c-1]
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"
    return border, align

# Status colorization
STATUS_FILL = {
    "Built":             ("DCEDC8", "1B5E20"),  # light green / dark green text
    "Placeholder":       ("FFE0B2", "5D4037"),  # light amber / brown
    "TBD":               ("FFCDD2", "B71C1C"),  # light red / dark red
    "Locked":            ("E1BEE7", "4A148C"),  # light purple
    "Reference":         ("E0E0E0", "424242"),  # light grey
    "Planned":           ("F5F5F5", "757575"),
    "Archive candidate": ("FFF9C4", "F57F17"),
}

def write_rows(ws, rows):
    border, align = style_sheet(ws)
    link_font = Font(color="C0392B", underline="single")
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = align
            cell.border = border
            # Hyperlinks: col 5 (live link), col 10 (vk equiv)
            if c in (5, 10) and val and isinstance(val, str) and val.startswith("http"):
                cell.hyperlink = val
                cell.font = link_font
            # Hyperlinks within "Label — URL" cells (cols 8, 9)
            elif c in (8, 9) and val and isinstance(val, str) and " — http" in val:
                u = val.split(" — ", 1)[1]
                cell.hyperlink = u
                cell.font = link_font
            # Status color (col 4)
            if c == 4 and val in STATUS_FILL:
                bg, fg = STATUS_FILL[val]
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Calibri", size=11, bold=True, color=fg)
                cell.alignment = Alignment(vertical="center", horizontal="center")
        ws.row_dimensions[r].height = 60

def main():
    wb = Workbook()
    wb.remove(wb.active)
    sheets = [
        ("T1 Editorial Light",      render_items(T1_ITEMS, "T1 Editorial Light")),
        ("T2 Deep Article (HOF)",   render_items(T2_ITEMS, "T2 Deep Article")),
        ("T3 Program & Service",    render_items(T3_ITEMS, "T3 Program/Service")),
        ("T4 Promo & Drop",         render_items(T4_ITEMS, "T4 Promo/Drop")),
        ("T5 Policy & Long-form",   render_items(T5_ITEMS, "T5 Policy/Long-form")),
        ("PLP variants (C1-C4)",    tab_plp_variants()),
        ("Core ecommerce (locked)", tab_core_ecommerce()),
        ("Reference & tools",       tab_reference_tools()),
        ("Archive",                 tab_archive()),
    ]
    for name, rows in sheets:
        ws = wb.create_sheet(name)
        write_rows(ws, rows)
    out = ROOT / "vk-resources-inventory.xlsx"
    wb.save(out)
    total = sum(len(r) for _, r in sheets)
    print(f"wrote {out}")
    print(f"  sheets: {len(sheets)}  rows: {total}")
    for name, rows in sheets:
        print(f"  - {name}: {len(rows)} rows")

if __name__ == "__main__":
    main()
